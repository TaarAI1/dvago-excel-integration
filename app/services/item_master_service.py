"""
Item Master import service.

Processing pipeline per row:
  1. Auto-detect header row (scan first 5 rows for known column names)
  2. Parse rows into dicts keyed by column name
  3. Authenticate RetailPro ONCE per batch — cache Auth-Session
  4. For each row:
      a. Check/create DCS   (Oracle lookup, then RetailPro API) — dedup by DCS_CODE
      b. Check/create Vendor (Oracle lookup, then RetailPro API) — dedup by VEND_CODE
      c. Fetch taxcodesid, sbssid from Oracle                    — dedup by code/no
      d. GET inventory by UPC → update if exists, create if not
      e. POST /api/backoffice/inventory?action=InventorySaveItems

Caches are shared across all rows in a batch so each unique DCS/Vendor/tax
code/subsidiary is only looked up once.
"""
import io
import logging
import uuid as _uuid_mod
from typing import Any, Optional

from app.core.timezone import now_pkt

import httpx
import openpyxl

logger = logging.getLogger(__name__)

# ── In-memory cancel state (same pattern as sales_export_job) ──────────────────
# run_id of the currently executing manual import (None when idle)
_active_import_id: Optional[str] = None
_cancel_requests: set = set()


def get_active_import_id() -> Optional[str]:
    return _active_import_id


def request_cancel_import() -> bool:
    """Signal the running import to stop after the current row finishes."""
    global _active_import_id
    if not _active_import_id:
        return False
    _cancel_requests.add(_active_import_id)
    logger.info("Cancel requested for item master import %s", _active_import_id)
    return True


def _is_import_cancelled(import_id: str) -> bool:
    return import_id in _cancel_requests

# ── Column → JSON field maps ────────────────────────────────────────────────────

# Main inventory object fields
MAIN_FIELD_MAP: dict[str, str] = {
    "SID": "sid",
    "CREATED_BY": "createdby",
    "CREATED_DATETIME": "createddatetime",
    "MODIFIED_BY": "modifiedby",
    "MODIFIED_DATETIME": "modifieddatetime",
    "CONTROLLER_SID": "controllersid",
    "ORIGIN_APPLICATION": "originapplication",
    "POST_DATE": "postdate",
    # "ROW_VERSION": "rowversion",  # excluded from payload
    "TENANT_SID": "tenantsid",
    "INVN_ITEM_UID": "invnitemuid",
    "SBS_SID": "sbssid",
    "ALU": "alu",
    "STYLE_SID": "stylesid",
    "DCS_SID": "dcssid",
    "VEND_SID": "vendsid",
    "DESCRIPTION1": "description1",
    "DESCRIPTION2": "description2",
    "DESCRIPTION3": "description3",
    "DESCRIPTION4": "description4",
    "ATTRIBUTE": "attribute",
    "COST": "cost",
    "SPIF": "spif",
    "CURRENCY_SID": "currencysid",
    "LAST_SOLD_DATE": "lastsolddate",
    "MARKDOWN_DATE": "markdowndate",
    "DISCONTINUED_DATE": "discontinueddate",
    "TAX_CODE_SID": "taxcodesid",
    "UDF1_FLOAT": "udf1float",
    "UDF2_FLOAT": "udf2float",
    "UDF3_FLOAT": "udf3float",
    "UDF1_DATE": "udf1date",
    "UDF2_DATE": "udf2date",
    "UDF3_DATE": "udf3date",
    "ITEM_SIZE": "itemsize",
    "FC_COST": "fccost",
    "FIRST_RCVD_DATE": "firstrcvddate",
    "LAST_RCVD_DATE": "lastrcvddate",
    "COMM_SID": "commsid",
    "DISC_SCHEDULE_SID": "discschedulesid",
    "UDF1_STRING": "udf1string",
    "UDF2_STRING": "udf2string",
    "UDF3_STRING": "udf3string",
    "UDF4_STRING": "udf4string",
    "UDF5_STRING": "udf5string",
    "SELLABLE_DATE": "sellabledate",
    "ORDERABLE_DATE": "orderabledate",
    "USE_QTY_DECIMALS": "useqtydecimals",
    "FORCE_ORIG_TAX": "forceorigtax",
    "FST_PRICE": "fstprice",
    "DESCRIPTION": "description",
    "REGIONAL": "regional",
    "ACTIVE": "active",
    "QTY_PER_CASE": "qtypercase",
    "UPC": "upc",
    "MAX_DISC_PERC1": "maxdiscperc1",
    "MAX_DISC_PERC2": "maxdiscperc2",
    "ITEM_NO": "itemno",
    "SERIAL_TYPE": "serialtype",
    "LOT_TYPE": "lottype",
    "KIT_TYPE": "kittype",
    "SCALE_SID": "scalesid",
    "PROMO_QTYDISCWEIGHT": "promoqtydiscweight",
    "PROMO_INVENEXCLUDE": "promoinvenexclude",
    "LAST_RCVD_COST": "lastrcvdcost",
    "NON_INVENTORY": "noninventory",
    "NON_COMMITED": "noncommitted",
    "ORDERABLE": "orderable",
    "LTY_PRICE_IN_POINTS": "ltypriceinpoints",
    "LTY_POINTS_EARNED": "ltypointsearned",
    "ITEM_STATE": "itemstate",
    "PUBLISH_STATUS": "publishstatus",
    "MIN_ORD_QTY": "minordqty",
    "VENDOR_LIST_COST": "vendorlistcost",
    "TRADE_DISC_PERC": "tradediscpercent",
    "LONG_DESCRIPTION": "longdescription",
    "TEXT1": "text1",
    "TEXT2": "text2",
    "TEXT3": "text3",
    "TEXT4": "text4",
    "TEXT5": "text5",
    "TEXT6": "text6",
    "TEXT7": "text7",
    "TEXT8": "text8",
    "TEXT9": "text9",
    "TEXT10": "text10",
    "D_NAME": "dname",
    "S_NAME": "sname",
    "C_NAME": "cname",
    "VEND_NAME": "vendorname",
    "TAX_NAME": "taxname",
    "TAX_CODE": "taxcode",
}

# invnextend sub-object fields
INVNEXTEND_MAP: dict[str, str] = {
    "UDF6_STRING": "udf6string",
    "UDF7_STRING": "udf7string",
    "UDF8_STRING": "udf8string",
    "UDF9_STRING": "udf9string",
    "UDF10_STRING": "udf10string",
    "UDF11_STRING": "udf11string",
    "UDF12_STRING": "udf12string",
    "UDF13_STRING": "udf13string",
    "UDF14_STRING": "udf14string",
    "UDF15_STRING": "udf15string",
    "UDF1_LARGE_STRING": "udf1largestring",
    "UDF2_LARGE_STRING": "udf2largestring",
}

# ── Field-type schema (from RetailPro API data-type definitions) ───────────────
# Maps each JSON payload field name to its required Python type category.
# Used by build_payload to coerce every value before sending.
# Omitted fields are left as-is (string / datetime handled by _to_json).
FIELD_TYPES: dict[str, str] = {
    # ── int64 ──────────────────────────────────────────────────────────────────
    "sid":                  "int",
    "controllersid":        "int",
    "tenantsid":            "int",
    "invnitemuid":          "int",
    "sbssid":               "int",
    "stylesid":             "int",
    "dcssid":               "int",
    "vendsid":              "int",
    "currencysid":          "int",
    "taxcodesid":           "int",
    "commsid":              "int",
    "discschedulesid":      "int",
    "scalesid":             "int",
    "upc":                  "int",
    "activestoresid":       "int",
    "activepricelevelsid":  "int",
    "activeseasonsid":      "int",
    "docitemsid":           "int",
    # ── int32 ──────────────────────────────────────────────────────────────────
    "rowversion":           "int",
    "udf1float":            "int",
    "udf2float":            "int",
    "udf3float":            "int",
    "itemno":               "int",
    "vendorid":             "int",
    "scaleno":              "int",
    # ── int16 ──────────────────────────────────────────────────────────────────
    "useqtydecimals":       "int",
    "serialtype":           "int",
    "lottype":              "int",
    "kittype":              "int",
    "itemstate":            "int",
    "publishstatus":        "int",
    "sbsno":                "int",
    # ── float ──────────────────────────────────────────────────────────────────
    "cost":                 "float",
    "spif":                 "float",
    "fccost":               "float",
    "fstprice":             "float",
    "lastrcvdcost":         "float",
    "qtypercase":           "float",
    "maxdiscperc1":         "float",
    "maxdiscperc2":         "float",
    "promoqtydiscweight":   "float",
    "ltypriceinpoints":     "float",
    "ltypointsearned":      "float",
    "minordqty":            "float",
    "vendorlistcost":       "float",
    "tradediscpercent":     "float",
    "height":               "float",
    "length":               "float",
    "width":                "float",
    "docqty":               "float",
    "doccaseqty":           "float",
    "docprice":             "float",
    "doccost":              "float",
    # ── boolean ────────────────────────────────────────────────────────────────
    "orderable":            "bool",
    "regional":             "bool",
    "active":               "bool",
    "promoinvenexclude":    "bool",
    "noninventory":         "bool",
    "noncommitted":         "bool",
    "forceorigtax":         "bool",
    "specialorder":         "bool",
    # ── int64 (invnextend) ────────────────────────────────────────────────────
    "invnsbsitemsid":           "int",
    # ── float (active-store computed) ────────────────────────────────────────
    "actstrdbprice":            "float",
    "actstrprice":              "float",
    "actstrpricewt":            "float",
    "actstrohqty":              "float",
    "actstrcaseqty":            "float",
    "actstravailqty":           "float",
    "actstrextcost":            "float",
    "actstrextprice":           "float",
    "actstrextpricewt":         "float",
    "actstrtaxpctg":            "float",
    "actstrtaxamt":             "float",
    "actstrtaxpctg2":           "float",
    "actstrtaxamt2":            "float",
    "actstrexttaxamt":          "float",
    "actstrexttaxamt1":         "float",
    "actstrexttaxamt2":         "float",
    "actstrmarginpctg":         "float",
    "actstrmarginamt":          "float",
    "actstrextmarginamt":       "float",
    "actstrmarginamtwt":        "float",
    "actstrextmarginamtwt":     "float",
    "actstrmarkuppctg":         "float",
    "actstrcoefficient":        "float",
    "actstrminqty":             "float",
    "actstrminextcost":         "float",
    "actstrminextprice":        "float",
    "actstrminextpricewt":      "float",
    "actstrmaxqty":             "float",
    "actstrmaxextcost":         "float",
    "actstrmaxextprice":        "float",
    "actstrmaxextpricewt":      "float",
    # ── float (comparative-store computed) ───────────────────────────────────
    "cmpstrohqty":              "float",
    "cmpstrextcost":            "float",
    "cmpstrextprice":           "float",
    "cmpstrminqty":             "float",
    "cmpstrminextcost":         "float",
    "cmpstrminextprice":        "float",
    "cmpstrminextpricewt":      "float",
    "cmpstrmaxqty":             "float",
    "cmpstrmaxextcost":         "float",
    "cmpstrmaxextprice":        "float",
    "cmpstrmaxextpricewt":      "float",
}

# All known column names used for header-row auto-detection
_KNOWN_COLUMNS: set[str] = (
    set(MAIN_FIELD_MAP)
    | set(INVNEXTEND_MAP)
    | {"DCS_CODE", "VEND_CODE", "SBS_NO"}
)

# ── Excel parsing ───────────────────────────────────────────────────────────────

def _norm(value: Any) -> str:
    """Uppercase + strip a cell value for comparison."""
    return str(value).strip().upper() if value is not None else ""


def detect_header_row(ws) -> int:
    """
    Scan the first 5 rows and return the 1-based index of the row whose
    cells best match the known column set (at least 3 hits required).
    """
    best_row, best_score = 1, 0
    for row_idx in range(1, 6):
        cells = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        score = sum(1 for v in cells if _norm(v) in _KNOWN_COLUMNS)
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score < 3:
        raise ValueError(
            "Could not find a header row in the first 5 rows. "
            "Expected at least 3 matching column names (e.g. UPC, DESCRIPTION1, COST)."
        )
    return best_row


def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse Excel bytes into a list of row dicts.

    Mapping is strictly by column name (case-insensitive, whitespace-stripped),
    so column order in the file does not matter.  Columns whose header cell is
    blank or None are silently ignored.  Rows without a UPC value are skipped.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    hdr_idx = detect_header_row(ws)

    # column-index → normalised column name (blank-header columns excluded)
    col_map: dict[int, str] = {}
    rows_out: list[dict] = []

    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=hdr_idx, values_only=True), start=hdr_idx
    ):
        if row_idx == hdr_idx:
            col_map = {
                i: _norm(v)
                for i, v in enumerate(row_vals)
                if _norm(v)          # skip empty / None header cells
            }
            continue
        # Build row dict by column name; columns not present in the header are
        # ignored, so extra trailing cells never cause an IndexError.
        rd = {col_map[i]: v for i, v in enumerate(row_vals) if i in col_map}
        upc_raw = rd.get("UPC")
        upc_str = str(upc_raw).strip() if upc_raw else ""
        # Skip completely blank rows (no UPC and no description at all)
        desc = str(rd.get("DESCRIPTION1") or rd.get("DESCRIPTION") or "").strip()
        if not upc_str and not desc:
            continue
        rd["UPC"] = upc_str          # "" when absent — payload builder will omit it
        rows_out.append(rd)

    wb.close()
    return rows_out


# ── Oracle helpers ──────────────────────────────────────────────────────────────

async def _oracle_scalar(sql: str, oc: dict) -> Optional[str]:
    """Run a query and return the first cell of the first row, or None."""
    from app.services.oracle_service import run_query
    df = await run_query(
        oc["host"], oc["port"], oc["service_name"], oc["username"], oc["password"], sql
    )
    if df is None or df.is_empty():
        return None
    return str(df.row(0)[0])


# ── Cache-aware SID resolvers ───────────────────────────────────────────────────

async def _sbssid(sbs_no: Any, cache: dict, oc: dict) -> Optional[str]:
    key = str(sbs_no).strip() if sbs_no else "1"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            f"SELECT sid FROM RPS.SUBSIDIARY WHERE sbs_no = {key}", oc
        )
    return cache[key]


async def _dcssid(
    dcs_code: str, row: dict,
    dcs_cache: dict, sbs_cache: dict,
    oc: dict, http: httpx.AsyncClient,
    base_url: str, auth_session: str,
    debug: Optional[dict] = None,
) -> Optional[str]:
    if not dcs_code:
        if debug is not None:
            debug.update({"dcs_code": "(empty)", "skipped": "DCS_CODE column missing or empty"})
        return None
    if dcs_code in dcs_cache:
        if debug is not None:
            debug.update({"dcs_code": dcs_code, "source": "cache", "final_sid": dcs_cache[dcs_code]})
        return dcs_cache[dcs_code]

    sid = await _oracle_scalar(
        f"SELECT sid FROM rps.dcs WHERE dcs_code = '{dcs_code}'", oc
    )
    if debug is not None:
        debug.update({"dcs_code": dcs_code, "oracle_found": bool(sid), "oracle_sid": sid})

    if not sid:
        sbssid_val = await _sbssid(row.get("SBS_NO", "1"), sbs_cache, oc)
        # d/c/s are derived by splitting dcs_code into 3-char chunks
        d_part = dcs_code[0:3]
        c_part = dcs_code[3:6]
        s_part = dcs_code[6:9]
        body = {k: v for k, v in {
            "originapplication": "RProPrismWeb",
            "active": 1,
            "sbssid": sbssid_val,
            "regional": False,
            "d": d_part,
            "c": c_part,
            "s": s_part,
            "dcscode": dcs_code,
            "dname": row.get("D_NAME") or "",
            "cname": row.get("C_NAME") or "",
            "sname": row.get("S_NAME") or "",
        }.items() if v is not None}
        resp = await http.post(
            f"{base_url}/api/backoffice/dcs",
            json={"data": [body]},
            headers={"Auth-Session": auth_session, "accept": "application/json, version=2",
                     "Content-Type": "application/json"},
        )
        try:
            resp_json = resp.json()
        except Exception:
            resp_json = {"raw": resp.text}
        if debug is not None:
            debug.update({
                "api_payload": {"data": [body]},
                "api_status": resp.status_code,
                "api_response": resp_json,
            })
        api_data = (resp_json.get("data") or []) if isinstance(resp_json, dict) else []
        sid = api_data[0].get("sid") if api_data else None
        if not sid:
            # Retry Oracle after create
            sid = await _oracle_scalar(
                f"SELECT sid FROM rps.dcs WHERE dcs_code = '{dcs_code}'", oc
            )
            if debug is not None:
                debug["oracle_retry_sid"] = sid

    if debug is not None:
        debug["final_sid"] = sid
    dcs_cache[dcs_code] = sid
    return sid


async def _vendsid(
    vend_code: str, row: dict,
    vend_cache: dict, sbs_cache: dict,
    oc: dict, http: httpx.AsyncClient,
    base_url: str, auth_session: str,
    debug: Optional[dict] = None,
) -> Optional[str]:
    if not vend_code:
        if debug is not None:
            debug.update({"vend_code": "(empty)", "skipped": "VEND_CODE column missing or empty"})
        return None
    if vend_code in vend_cache:
        if debug is not None:
            debug.update({"vend_code": vend_code, "source": "cache", "final_sid": vend_cache[vend_code]})
        return vend_cache[vend_code]

    sid = await _oracle_scalar(
        f"SELECT sid FROM rps.vendor WHERE vend_code = '{vend_code}'", oc
    )
    if debug is not None:
        debug.update({"vend_code": vend_code, "oracle_found": bool(sid), "oracle_sid": sid})

    if not sid:
        sbssid_val = await _sbssid(row.get("SBS_NO", "1"), sbs_cache, oc)
        body = {k: v for k, v in {
            "originapplication": "RProPrismWeb",
            "sbssid": sbssid_val,
            "regional": False,
            "vendcode": vend_code,
            "vendname": row.get("VEND_NAME") or vend_code,
            "active": True,
        }.items() if v is not None}
        resp = await http.post(
            f"{base_url}/api/backoffice/vendor",
            json={"data": [body]},
            headers={"Auth-Session": auth_session, "accept": "application/json, version=2",
                     "Content-Type": "application/json"},
        )
        try:
            resp_json = resp.json()
        except Exception:
            resp_json = {"raw": resp.text}
        if debug is not None:
            debug.update({
                "api_payload": {"data": [body]},
                "api_status": resp.status_code,
                "api_response": resp_json,
            })
        api_data = (resp_json.get("data") or []) if isinstance(resp_json, dict) else []
        sid = api_data[0].get("sid") if api_data else None
        if not sid:
            sid = await _oracle_scalar(
                f"SELECT sid FROM rps.vendor WHERE vend_code = '{vend_code}'", oc
            )
            if debug is not None:
                debug["oracle_retry_sid"] = sid

    if debug is not None:
        debug["final_sid"] = sid
    vend_cache[vend_code] = sid
    return sid


async def _taxcodesid(tax_code: str, tax_cache: dict, oc: dict) -> Optional[str]:
    if not tax_code:
        return None
    if tax_code not in tax_cache:
        tax_cache[tax_code] = await _oracle_scalar(
            f"SELECT sid FROM rps.tax_code WHERE tax_code = '{tax_code}'", oc
        )
    return tax_cache[tax_code]


async def _pref_reason_sid(cache: dict, oc: dict) -> Optional[str]:
    """Return the SID for the 'MANUALLY' preferred reason — cached for the whole batch."""
    key = "__MANUALLY__"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            "SELECT sid FROM RPS.PREF_REASON WHERE name = 'MANUALLY'", oc
        )
    return cache[key]


async def _store_sid(store_no: Any, sbssid: Optional[str], cache: dict, oc: dict) -> Optional[str]:
    """Return the store SID for the given store_no + subsidiary SID pair."""
    sn = str(store_no).strip() if store_no else "1"
    if not sbssid:
        return None
    key = f"{sn}::{sbssid}"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            f"SELECT sid FROM rps.store WHERE store_no = {sn} AND sbs_sid = '{sbssid}'", oc
        )
    return cache[key]


# ── Payload builder ─────────────────────────────────────────────────────────────

def _to_json(v: Any) -> Any:
    import re as _re
    from datetime import datetime as _dt
    if v is None:
        return None
    if isinstance(v, _dt):
        return v.isoformat()
    # Float whole-numbers (e.g. 1e12 read from Excel) → "1000000000000"
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return v
    # String scientific-notation (e.g. "1E+12") → "1000000000000"
    if isinstance(v, str):
        stripped = v.strip()
        if _re.match(r'^-?\d+\.?\d*[eE][+\-]?\d+$', stripped):
            try:
                num = float(stripped)
                if num == int(num):
                    return str(int(num))
                return str(num)
            except (ValueError, OverflowError):
                pass
        # Strip single quotes: RetailPro's Delphi-based server uses ' as its
        # native string delimiter, so a literal ' inside a JSON string value
        # confuses its tokeniser and causes adjacent integer fields to be read
        # as '' → "'' is not a valid integer value" serialiser error.
        return stripped.replace("'", "")
    return v


def build_payload(
    row: dict,
    sid_overrides: dict,
    existing_item: Optional[dict],
    pref_reason_sid: Optional[str] = None,
    store_sid: Optional[str] = None,
) -> dict:
    """
    Build the RetailPro InventorySaveItems payload for one row.

    The outer shape expected by the API is:
        {"data": [<this return value>]}

    where this return value is:
        {
            "OriginApplication": "RProPrismWeb",
            "PrimaryItemDefinition": { dcssid, vendsid, description1, … , sid },
            "InventoryItems": [ { …all fields…, "invnextend": […] } ],
            "SavingStyle": false,
            "localdate": "YYYY-MM-DDTHH:MM:SS",
            "DefaultReasonSidForQtyMemo":   <pref_reason_sid>,
            "DefaultReasonSidForCostMemo":  <pref_reason_sid>,
            "DefaultReasonSidForPriceMemo": <pref_reason_sid>,
            …
        }

    sid_overrides:    {json_field: value} – computed SIDs (dcssid, vendsid, etc.)
    existing_item:    API response item dict for update; None for create.
    pref_reason_sid:  SID from RPS.PREF_REASON WHERE name = 'MANUALLY'
    store_sid:        SID from rps.store for the row's store_no + subsidiary
    """
    # ── InventoryItems entry ──────────────────────────────────────────────────
    inv_item: dict = {}

    if existing_item:
        inv_item["sid"] = existing_item.get("sid")
    else:
        inv_item["sid"] = None

    # Map Excel columns → JSON fields (skip None and empty strings)
    for col, json_key in MAIN_FIELD_MAP.items():
        if col in row and row[col] is not None and str(row[col]).strip() != "" and json_key not in inv_item:
            inv_item[json_key] = _to_json(row[col])

    # Inject computed SIDs (overwrite whatever came from Excel)
    inv_item.update(sid_overrides)

    # activestoresid — resolved from Oracle; set before coercion so it is typed correctly
    if store_sid:
        inv_item["activestoresid"] = store_sid

    # Coerce every field to its declared RetailPro API data type.
    # Empty strings are treated the same as None (nulled out) so they are never
    # sent as "" for an integer/float/bool field — that would cause the RetailPro
    # "'' is not a valid integer value" serialiser error.
    for _key, _dtype in FIELD_TYPES.items():
        if _key not in inv_item or inv_item[_key] is None:
            continue
        raw = inv_item[_key]
        if isinstance(raw, str) and raw.strip() == "":
            inv_item[_key] = None
            continue
        try:
            if _dtype == "int":
                # Avoid int(float(x)): float64 only has ~15 significant digits,
                # which silently truncates int64 SIDs (18-19 digits).
                inv_item[_key] = int(str(raw).split(".")[0])
            elif _dtype == "float":
                inv_item[_key] = float(raw)
            elif _dtype == "bool":
                inv_item[_key] = bool(int(float(raw)))
        except (ValueError, TypeError):
            inv_item[_key] = None

    # ── Shared SID coercion helper ────────────────────────────────────────────
    def _coerce_sid(val: Any) -> Optional[int]:
        """Safely coerce a SID string/number to int without float precision loss.
        Returns None for None, empty string, or non-numeric values."""
        if val is None:
            return None
        s = str(val).strip()
        if not s or s.lower() == "none":
            return None
        try:
            return int(s.split(".")[0])
        except (ValueError, TypeError):
            return None

    # ── invnextend sub-object ─────────────────────────────────────────────────
    extend: dict = {}
    if existing_item:
        ex_ext = (existing_item.get("invnextend") or [{}])[0]
        # Coerce integer fields from the API response — they can come back as ""
        # which would fail the RetailPro serialiser on the next save.
        ext_sid = _coerce_sid(ex_ext.get("sid"))
        if ext_sid is not None:
            extend["sid"] = ext_sid
        ext_rv = _coerce_sid(ex_ext.get("rowversion"))
        if ext_rv is not None:
            extend["rowversion"] = ext_rv
        ext_parent = _coerce_sid(existing_item.get("sid"))
        if ext_parent is not None:
            extend["invnsbsitemsid"] = ext_parent

    for col, json_key in INVNEXTEND_MAP.items():
        if col in row and row[col] is not None and str(row[col]).strip() != "":
            extend[json_key] = _to_json(row[col])

    inv_item["invnextend"] = [extend]

    # ── PrimaryItemDefinition ─────────────────────────────────────────────────
    # Key order: sid → dcssid → vendsid → description1 → description2 → attribute → itemsize
    # dcssid and vendsid are always included (can be null, same as sid).

    primary_def: dict = {
        "sid":    _coerce_sid(existing_item.get("sid")) if existing_item else None,
        "dcssid": _coerce_sid(sid_overrides.get("dcssid")),
        "vendsid": _coerce_sid(sid_overrides.get("vendsid")),
    }
    for k, v in {
        "description1": inv_item.get("description1"),
        "description2": inv_item.get("description2"),
        "attribute":    inv_item.get("attribute"),
        "itemsize":     inv_item.get("itemsize"),
    }.items():
        if v is not None and str(v).strip() != "":
            primary_def[k] = v

    # ── Outer wrapper (OriginApplication + PrimaryItemDefinition always first) ─
    outer: dict = {
        "OriginApplication":     "RProPrismWeb",
        "PrimaryItemDefinition": primary_def,
        "InventoryItems":        [inv_item],
        "UpdateStyleDefinition": False,
        "UpdateStyleCost":       False,
        "UpdateStylePrice":      False,
    }
    if pref_reason_sid:
        reason_int = _coerce_sid(pref_reason_sid)
        outer["DefaultReasonSidForQtyMemo"]   = reason_int
        outer["DefaultReasonSidForCostMemo"]  = reason_int
        outer["DefaultReasonSidForPriceMemo"] = reason_int
    return outer


# ── Per-row processor ───────────────────────────────────────────────────────────

async def process_row(
    row: dict,
    auth_session: str,
    oc: dict,
    base_url: str,
    http: httpx.AsyncClient,
    dcs_cache: dict,
    vend_cache: dict,
    tax_cache: dict,
    sbs_cache: dict,
    pref_reason_cache: dict,
    store_cache: dict,
) -> dict:
    upc = str(row.get("UPC") or "").strip()
    has_upc = bool(upc)
    result: dict = {"upc": upc, "action": None, "sid": None, "ok": False, "error": None,
                    "description": row.get("DESCRIPTION1") or row.get("DESCRIPTION") or ""}
    dcs_debug: dict = {}
    vend_debug: dict = {}
    try:
        dcs_code  = str(row.get("DCS_CODE")  or "").strip()
        vend_code = str(row.get("VEND_CODE") or "").strip()
        tax_code  = str(row.get("TAX_CODE")  or "").strip()
        sbs_no    = str(row.get("SBS_NO")    or "1").strip()

        rp_headers = {
            "Auth-Session": auth_session,
            "accept": "application/json, version=2",
            "Content-Type": "application/json",
        }

        # ── 1. DCS ───────────────────────────────────────────────────────────
        dcssid_val = await _dcssid(
            dcs_code, row, dcs_cache, sbs_cache, oc, http, base_url, auth_session,
            debug=dcs_debug,
        )

        # ── 2. Vendor ────────────────────────────────────────────────────────
        vendsid_val = await _vendsid(
            vend_code, row, vend_cache, sbs_cache, oc, http, base_url, auth_session,
            debug=vend_debug,
        )

        # ── 3. Tax + subsidiary ──────────────────────────────────────────────
        taxcodesid_val = await _taxcodesid(tax_code, tax_cache, oc)
        sbssid_val     = await _sbssid(sbs_no, sbs_cache, oc)

        # ── 4. Preferred-reason SID (DefaultReasonSid* fields) ───────────────
        pref_reason_sid_val = await _pref_reason_sid(pref_reason_cache, oc)

        # ── 5. Store SID (activestoresid) ────────────────────────────────────
        store_no_val    = str(row.get("STORE_NO") or "1").strip()
        store_sid_val   = await _store_sid(store_no_val, sbssid_val, store_cache, oc)

        # ── 6. Check item by UPC (skip when UPC is empty — always create) ────
        if has_upc:
            check_resp = await http.get(
                f"{base_url}/api/backoffice/inventory",
                params={"filter": f'(upc,eq,"{upc}")', "cols": "*,invnextend.*"},
                headers=rp_headers,
            )
            check_data = check_resp.json().get("data") or []
        else:
            # No UPC supplied — RetailPro will assign one; always create
            check_data = []

        # dcssid and vendsid are always included (null is valid and expected in both
        # PrimaryItemDefinition and InventoryItems).  taxcodesid / sbssid are only
        # included when resolved because sbssid has a per-update fallback below.
        sid_overrides: dict = {
            "dcssid":  dcssid_val,   # always present, may be None
            "vendsid": vendsid_val,  # always present, may be None
        }
        if taxcodesid_val is not None:
            sid_overrides["taxcodesid"] = taxcodesid_val
        if sbssid_val is not None:
            sid_overrides["sbssid"] = sbssid_val

        if len(check_data) > 0:
            existing = check_data[0]
            # For update, preserve existing sbssid if we couldn't resolve it
            if "sbssid" not in sid_overrides:
                sid_overrides["sbssid"] = existing.get("sbssid")
            payload = build_payload(
                row, sid_overrides, existing_item=existing,
                pref_reason_sid=pref_reason_sid_val,
                store_sid=store_sid_val,
            )
            action = "updated"
        else:
            payload = build_payload(
                row, sid_overrides, existing_item=None,
                pref_reason_sid=pref_reason_sid_val,
                store_sid=store_sid_val,
            )
            action = "created"

        # ── 7. Upsert ────────────────────────────────────────────────────────
        save_resp = await http.post(
            f"{base_url}/api/backoffice/inventory",
            params={"action": "InventorySaveItems"},
            json={"data": [payload]},
            headers=rp_headers,
        )
        save_json = save_resp.json()
        saved_data = save_json.get("data") or []
        api_errors = save_json.get("errors")

        # Success: errors is null AND data array is non-empty
        ok = (api_errors is None) and (len(saved_data) > 0)

        # Extract SID and UPC from the response
        sid = None
        resp_upc: Optional[str] = None
        if saved_data:
            inv_items = saved_data[0].get("inventoryitems") or []
            if inv_items:
                sid      = inv_items[0].get("sid")
                resp_upc = inv_items[0].get("upc") or None
            if not sid:
                sid = saved_data[0].get("newstylesid")

        # When UPC was empty use whatever RetailPro assigned in the response
        final_upc = upc or resp_upc or ""

        if ok:
            result.update({
                "action": action, "sid": sid, "ok": True,
                "upc": final_upc,
                "resp_upc": resp_upc,
                "api_response": save_json,
                "_dcs_debug": dcs_debug,
                "_vend_debug": vend_debug,
            })
        else:
            result.update({
                "action": action, "ok": False,
                "api_response": save_json,
                "error": save_resp.text,
                "_payload_sent": {"data": [payload]},
                "_dcs_debug": dcs_debug,
                "_vend_debug": vend_debug,
            })

    except Exception as exc:
        result["error"] = str(exc)
        result["_dcs_debug"] = dcs_debug
        result["_vend_debug"] = vend_debug

    return result


# ── DB persistence ──────────────────────────────────────────────────────────────

async def _persist_result(row: dict, result: dict, source_file: str) -> None:
    """Save a processed row result to the documents table."""
    import uuid as _uuid
    from app.db.postgres import get_session
    from app.models.document import Document

    ok = result.get("ok", False)
    error_msg = result.get("error")

    # Store the serialisable row dict (convert datetime → str)
    safe_row = {k: (_to_json(v)) for k, v in row.items()}

    # When UPC was empty in the source file but RetailPro returned one, save it
    # so the grid can display it under the UPC column.
    if ok and result.get("resp_upc") and not safe_row.get("UPC"):
        safe_row["UPC"] = result["resp_upc"]

    import json as _json

    # On failure, embed the sent payload so the UI can show it alongside the error.
    # Stored as a JSON string (not a nested object) to preserve key insertion order
    # when retrieved from PostgreSQL JSONB.
    if not ok and result.get("_payload_sent"):
        safe_row["_payload_sent"] = _json.dumps(result["_payload_sent"], indent=2, ensure_ascii=False)

    # Always store DCS / vendor debug so the UI can show what happened to each SID.
    if result.get("_dcs_debug"):
        safe_row["_dcs_debug"] = _json.dumps(result["_dcs_debug"], indent=2, ensure_ascii=False)
    if result.get("_vend_debug"):
        safe_row["_vend_debug"] = _json.dumps(result["_vend_debug"], indent=2, ensure_ascii=False)

    async with get_session() as session:
        async with session.begin():
            doc = Document(
                id=_uuid.uuid4(),
                document_type="item_master",
                original_data=safe_row,
                retailprosid=result.get("sid") if ok else None,
                posted=ok,
                has_error=not ok,
                error_message=error_msg,          # full API response on failure
                source_file=source_file,
                posted_at=now_pkt() if ok else None,
            )
            session.add(doc)


# ── CSV parsing ─────────────────────────────────────────────────────────────────

def parse_csv_item_master(file_bytes: bytes) -> list[dict]:
    """
    Parse CSV bytes into item master row dicts — same format as parse_excel().
    Column headers are normalised to uppercase + stripped.
    Rows without a UPC value are skipped.
    Handles UTF-8 with or without BOM, and falls back to latin-1.
    """
    import csv as _csv
    import io as _io

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode CSV file (tried utf-8, latin-1).")

    reader = _csv.DictReader(_io.StringIO(text))
    rows_out: list[dict] = []
    for row in reader:
        normalised = {k.strip().upper(): v for k, v in row.items() if k}
        upc = str(normalised.get("UPC", "")).strip()
        # Skip completely blank rows (no UPC and no description)
        desc = str(normalised.get("DESCRIPTION1", "") or normalised.get("DESCRIPTION", "")).strip()
        if not upc and not desc:
            continue
        normalised["UPC"] = upc      # "" when absent — payload builder will omit it
        rows_out.append(normalised)
    return rows_out


# ── Shared batch runner ──────────────────────────────────────────────────────────

async def _run_rows_batch(rows: list[dict], source_file: str) -> dict:
    """
    Core pipeline shared by Excel and CSV paths:
      auth (once) → per-row processing with shared caches → persist each row.
    Supports cooperative cancellation via request_cancel_import().
    """
    from app.services.retailpro_auth import get_auth_session, sit_session, stand_session
    from app.db.settings_store import get_setting

    if not rows:
        return {"ok": False, "error": "No data rows found (all rows missing UPC).", "results": []}

    global _active_import_id
    import_id = str(_uuid_mod.uuid4())
    _active_import_id = import_id
    _cancel_requests.discard(import_id)

    oc = {
        "host":         (await get_setting("oracle_host"))         or "",
        "port":         int((await get_setting("oracle_port"))     or "1521"),
        "service_name": (await get_setting("oracle_service_name")) or "",
        "username":     (await get_setting("oracle_username"))     or "",
        "password":     (await get_setting("oracle_password"))     or "",
    }

    base_url     = ((await get_setting("retailpro_base_url")) or "").rstrip("/")
    auth_session = await get_auth_session()

    # Activate the session immediately after obtaining it
    await sit_session(base_url, auth_session)

    dcs_cache:          dict = {}
    vend_cache:         dict = {}
    tax_cache:          dict = {}
    sbs_cache:          dict = {}
    pref_reason_cache:  dict = {}
    store_cache:        dict = {}

    cancelled = False
    results: list[dict] = []
    try:
        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as http:
            for row in rows:
                # Check for kill signal before each row
                if _is_import_cancelled(import_id):
                    logger.info(
                        "Item master import %s cancelled — stopped after %d / %d rows",
                        import_id, len(results), len(rows),
                    )
                    cancelled = True
                    break

                row_result = await process_row(
                    row=row,
                    auth_session=auth_session,
                    oc=oc,
                    base_url=base_url,
                    http=http,
                    dcs_cache=dcs_cache,
                    vend_cache=vend_cache,
                    tax_cache=tax_cache,
                    sbs_cache=sbs_cache,
                    pref_reason_cache=pref_reason_cache,
                    store_cache=store_cache,
                )
                results.append(row_result)
                try:
                    await _persist_result(row, row_result, source_file)
                except Exception as db_exc:
                    logger.warning("DB persist failed for UPC %s: %s", row_result["upc"], db_exc)
                logger.info(
                    "Item %s: %s %s",
                    row_result["upc"],
                    row_result["action"],
                    "✓" if row_result["ok"] else f"✗ {row_result['error']}",
                )
    finally:
        # Always destroy the session and clear active state
        await stand_session(base_url, auth_session)
        if _active_import_id == import_id:
            _active_import_id = None
        _cancel_requests.discard(import_id)

    created = sum(1 for r in results if r["ok"] and r["action"] == "created")
    updated = sum(1 for r in results if r["ok"] and r["action"] == "updated")
    errors  = sum(1 for r in results if not r["ok"])

    return {
        "ok":        True,
        "cancelled": cancelled,
        "total":     len(results),
        "of_total":  len(rows),
        "created":   created,
        "updated":   updated,
        "errors":    errors,
        "results":   results,
    }


# ── Payload field diagnosis ──────────────────────────────────────────────────────

async def diagnose_bad_field(payload: dict) -> dict:
    """
    Binary-search through every field of InventoryItems[0] (and invnextend[0])
    to find the exact field that causes the RetailPro serialiser error
    "'' is not a valid integer value".

    Returns a dict like:
        {"found": True,  "section": "InventoryItems", "field": "itemsize", "value": "14'S"}
        {"found": False, "message": "..."}
    """
    from app.services.retailpro_auth import get_auth_session, sit_session, stand_session
    from app.db.settings_store import get_setting

    base_url     = ((await get_setting("retailpro_base_url")) or "").rstrip("/")
    auth_session = await get_auth_session()
    await sit_session(base_url, auth_session)

    rp_headers = {
        "Auth-Session": auth_session,
        "accept": "application/json, version=2",
        "Content-Type": "application/json",
    }

    def _is_int_error(resp_json: dict) -> bool:
        errors = resp_json.get("errors") or []
        return any(
            "is not a valid integer" in (e.get("errormsg") or "").lower()
            for e in errors
        )

    async def _send(test_inv_item: dict, test_extend: dict) -> bool:
        """Return True if this field combination triggers the integer error."""
        test_payload = {
            **payload,
            "InventoryItems": [{**test_inv_item, "invnextend": [test_extend]}],
        }
        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as http:
            resp = await http.post(
                f"{base_url}/api/backoffice/inventory",
                params={"action": "InventorySaveItems"},
                json={"data": [test_payload]},
                headers=rp_headers,
            )
        return _is_int_error(resp.json())

    try:
        orig_inv   = dict(payload.get("InventoryItems", [{}])[0])
        orig_ext   = dict((orig_inv.pop("invnextend", [{}]) or [{}])[0])

        # ── Reproduce the error with the full payload first ──────────────────
        if not await _send(orig_inv, orig_ext):
            return {"found": False, "message": "Error not reproducible — payload succeeds or returns a different error."}

        # ── Binary search on InventoryItems fields ───────────────────────────
        async def _bisect(pairs: list, section: str, base_inv: dict, base_ext: dict):
            if len(pairs) == 1:
                return {"found": True, "section": section, "field": pairs[0][0], "value": pairs[0][1]}
            mid   = len(pairs) // 2
            left  = pairs[:mid]
            right = pairs[mid:]
            for half in (left, right):
                test_inv = {**base_inv, **dict(half)} if section == "InventoryItems" else base_inv
                test_ext = {**base_ext, **dict(half)} if section == "invnextend"     else base_ext
                if await _send(test_inv, test_ext):
                    return await _bisect(half, section, base_inv, base_ext)
            # Both halves are clean individually — interaction between them.
            # Fall back to linear scan to find the first offending field.
            for k, v in pairs:
                test_inv = {**base_inv, k: v} if section == "InventoryItems" else base_inv
                test_ext = {**base_ext, k: v} if section == "invnextend"     else base_ext
                if await _send(test_inv, test_ext):
                    return {"found": True, "section": section, "field": k, "value": v}
            return {"found": False, "message": f"Could not isolate within {section} fields: {[k for k,_ in pairs]}"}

        # Start with a minimal baseline for InventoryItems: keep only sid
        base_inv = {"sid": orig_inv.get("sid")}
        inv_pairs = [(k, v) for k, v in orig_inv.items() if k != "sid"]

        result = await _bisect(inv_pairs, "InventoryItems", base_inv, orig_ext)
        if result.get("found"):
            return result

        # ── Binary search on invnextend fields ───────────────────────────────
        if orig_ext:
            ext_pairs = list(orig_ext.items())
            result = await _bisect(ext_pairs, "invnextend", orig_inv, {})
            if result.get("found"):
                return result

        return {"found": False, "message": "Error present but could not be attributed to a single field."}

    finally:
        await stand_session(base_url, auth_session)


# ── Batch entry points ───────────────────────────────────────────────────────────

async def process_excel_batch(file_bytes: bytes, source_file: str = "item_master_import.xlsx") -> dict:
    """Parse an Excel (.xlsx) file and run the full item master pipeline."""
    rows = parse_excel(file_bytes)
    return await _run_rows_batch(rows, source_file)


async def process_csv_batch(file_bytes: bytes, source_file: str = "item_master_import.csv") -> dict:
    """Parse a CSV file and run the full item master pipeline."""
    rows = parse_csv_item_master(file_bytes)
    return await _run_rows_batch(rows, source_file)
