"""
Item Master import service.

Processing pipeline per row:
  1. Auto-detect header row (scan first 5 rows for known column names)
  2. Parse rows into dicts keyed by column name
  3. Authenticate RetailPro ONCE per batch — cache Auth-Session
  4. For each row:
      a. Check/create DCS   (Oracle lookup, then RetailPro API) — dedup by DCS_CODE
      b. Check/create Vendor (Oracle lookup, then RetailPro API) — dedup by VEND_CODE
      c. Fetch taxcodesid, sbssid from Oracle                    — dedup by code/no
      d. GET inventory by UPC → update if exists, create if not
      e. POST /api/backoffice/inventory?action=InventorySaveItems

Caches are shared across all rows in a batch so each unique DCS/Vendor/tax
code/subsidiary is only looked up once.
"""
import io
import logging
from typing import Any, Optional

from app.core.timezone import now_pkt

import httpx
import openpyxl

logger = logging.getLogger(__name__)

# ── Column → JSON field maps ────────────────────────────────────────────────────

# Main inventory object fields
MAIN_FIELD_MAP: dict[str, str] = {
    "SID": "sid",
    "CREATED_BY": "createdby",
    "CREATED_DATETIME": "createddatetime",
    "MODIFIED_BY": "modifiedby",
    "MODIFIED_DATETIME": "modifieddatetime",
    "CONTROLLER_SID": "controllersid",
    "ORIGIN_APPLICATION": "originapplication",
    "POST_DATE": "postdate",
    # "ROW_VERSION": "rowversion",  # excluded from payload
    "TENANT_SID": "tenantsid",
    "INVN_ITEM_UID": "invnitemuid",
    "SBS_SID": "sbssid",
    "ALU": "alu",
    "STYLE_SID": "stylesid",
    "DCS_SID": "dcssid",
    "VEND_SID": "vendsid",
    "DESCRIPTION1": "description1",
    "DESCRIPTION2": "description2",
    "DESCRIPTION3": "description3",
    "DESCRIPTION4": "description4",
    "ATTRIBUTE": "attribute",
    "COST": "cost",
    "SPIF": "spif",
    "CURRENCY_SID": "currencysid",
    "LAST_SOLD_DATE": "lastsolddate",
    "MARKDOWN_DATE": "markdowndate",
    "DISCONTINUED_DATE": "discontinueddate",
    "TAX_CODE_SID": "taxcodesid",
    "UDF1_FLOAT": "udf1float",
    "UDF2_FLOAT": "udf2float",
    "UDF3_FLOAT": "udf3float",
    "UDF1_DATE": "udf1date",
    "UDF2_DATE": "udf2date",
    "UDF3_DATE": "udf3date",
    "ITEM_SIZE": "itemsize",
    "FC_COST": "fccost",
    "FIRST_RCVD_DATE": "firstrcvddate",
    "LAST_RCVD_DATE": "lastrcvddate",
    "COMM_SID": "commsid",
    "DISC_SCHEDULE_SID": "discschedulesid",
    "UDF1_STRING": "udf1string",
    "UDF2_STRING": "udf2string",
    "UDF3_STRING": "udf3string",
    "UDF4_STRING": "udf4string",
    "UDF5_STRING": "udf5string",
    "SELLABLE_DATE": "sellabledate",
    "ORDERABLE_DATE": "orderabledate",
    "USE_QTY_DECIMALS": "useqtydecimals",
    "FORCE_ORIG_TAX": "forceorigtax",
    "FST_PRICE": "fstprice",
    "DESCRIPTION": "description",
    "REGIONAL": "regional",
    "ACTIVE": "active",
    "QTY_PER_CASE": "qtypercase",
    "UPC": "upc",
    "MAX_DISC_PERC1": "maxdiscperc1",
    "MAX_DISC_PERC2": "maxdiscperc2",
    "ITEM_NO": "itemno",
    "SERIAL_TYPE": "serialtype",
    "LOT_TYPE": "lottype",
    "KIT_TYPE": "kittype",
    "SCALE_SID": "scalesid",
    "PROMO_QTYDISCWEIGHT": "promoqtydiscweight",
    "PROMO_INVENEXCLUDE": "promoinvenexclude",
    "LAST_RCVD_COST": "lastrcvdcost",
    "NON_INVENTORY": "noninventory",
    "NON_COMMITED": "noncommitted",
    "ORDERABLE": "orderable",
    "LTY_PRICE_IN_POINTS": "ltypriceinpoints",
    "LTY_POINTS_EARNED": "ltypointsearned",
    "ITEM_STATE": "itemstate",
    "PUBLISH_STATUS": "publishstatus",
    "MIN_ORD_QTY": "minordqty",
    "VENDOR_LIST_COST": "vendorlistcost",
    "TRADE_DISC_PERC": "tradediscpercent",
    "LONG_DESCRIPTION": "longdescription",
    "TEXT1": "text1",
    "TEXT2": "text2",
    "TEXT3": "text3",
    "TEXT4": "text4",
    "TEXT5": "text5",
    "TEXT6": "text6",
    "TEXT7": "text7",
    "TEXT8": "text8",
    "TEXT9": "text9",
    "TEXT10": "text10",
    "D_NAME": "dname",
    "S_NAME": "sname",
    "C_NAME": "cname",
    "VEND_NAME": "vendorname",
    "TAX_NAME": "taxname",
    "TAX_CODE": "taxcode",
}

# invnextend sub-object fields
INVNEXTEND_MAP: dict[str, str] = {
    "UDF6_STRING": "udf6string",
    "UDF7_STRING": "udf7string",
    "UDF8_STRING": "udf8string",
    "UDF9_STRING": "udf9string",
    "UDF10_STRING": "udf10string",
    "UDF11_STRING": "udf11string",
    "UDF12_STRING": "udf12string",
    "UDF13_STRING": "udf13string",
    "UDF14_STRING": "udf14string",
    "UDF15_STRING": "udf15string",
    "UDF1_LARGE_STRING": "udf1largestring",
    "UDF2_LARGE_STRING": "udf2largestring",
}

# All known column names used for header-row auto-detection
_KNOWN_COLUMNS: set[str] = (
    set(MAIN_FIELD_MAP)
    | set(INVNEXTEND_MAP)
    | {"DCS_CODE", "VEND_CODE", "SBS_NO"}
)

# ── Excel parsing ───────────────────────────────────────────────────────────────

def _norm(value: Any) -> str:
    """Uppercase + strip a cell value for comparison."""
    return str(value).strip().upper() if value is not None else ""


def detect_header_row(ws) -> int:
    """
    Scan the first 5 rows and return the 1-based index of the row whose
    cells best match the known column set (at least 3 hits required).
    """
    best_row, best_score = 1, 0
    for row_idx in range(1, 6):
        cells = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        score = sum(1 for v in cells if _norm(v) in _KNOWN_COLUMNS)
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score < 3:
        raise ValueError(
            "Could not find a header row in the first 5 rows. "
            "Expected at least 3 matching column names (e.g. UPC, DESCRIPTION1, COST)."
        )
    return best_row


def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse Excel bytes into a list of row dicts.
    Keys are normalised column names; values are raw cell values.
    Rows without a UPC are skipped.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    hdr_idx = detect_header_row(ws)

    headers: list[str] = []
    rows_out: list[dict] = []

    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=hdr_idx, values_only=True), start=hdr_idx
    ):
        if row_idx == hdr_idx:
            headers = [_norm(v) for v in row_vals]
            continue
        rd = {headers[i]: v for i, v in enumerate(row_vals) if i < len(headers)}
        upc = rd.get("UPC")
        if not upc:
            continue
        rd["UPC"] = str(upc).strip()
        rows_out.append(rd)

    wb.close()
    return rows_out


# ── Oracle helpers ──────────────────────────────────────────────────────────────

async def _oracle_scalar(sql: str, oc: dict) -> Optional[str]:
    """Run a query and return the first cell of the first row, or None."""
    from app.services.oracle_service import run_query
    df = await run_query(
        oc["host"], oc["port"], oc["service_name"], oc["username"], oc["password"], sql
    )
    if df is None or df.is_empty():
        return None
    return str(df.row(0)[0])


# ── Cache-aware SID resolvers ───────────────────────────────────────────────────

async def _sbssid(sbs_no: Any, cache: dict, oc: dict) -> Optional[str]:
    key = str(sbs_no).strip() if sbs_no else "1"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            f"SELECT sid FROM RPS.SUBSIDIARY WHERE sbs_no = {key}", oc
        )
    return cache[key]


async def _dcssid(
    dcs_code: str, row: dict,
    dcs_cache: dict, sbs_cache: dict,
    oc: dict, http: httpx.AsyncClient,
    base_url: str, auth_session: str,
    debug: Optional[dict] = None,
) -> Optional[str]:
    if not dcs_code:
        if debug is not None:
            debug.update({"dcs_code": "(empty)", "skipped": "DCS_CODE column missing or empty"})
        return None
    if dcs_code in dcs_cache:
        if debug is not None:
            debug.update({"dcs_code": dcs_code, "source": "cache", "final_sid": dcs_cache[dcs_code]})
        return dcs_cache[dcs_code]

    sid = await _oracle_scalar(
        f"SELECT sid FROM rps.dcs WHERE dcs_code = '{dcs_code}'", oc
    )
    if debug is not None:
        debug.update({"dcs_code": dcs_code, "oracle_found": bool(sid), "oracle_sid": sid})

    if not sid:
        sbssid_val = await _sbssid(row.get("SBS_NO", "1"), sbs_cache, oc)
        # d/c/s are derived by splitting dcs_code into 3-char chunks
        d_part = dcs_code[0:3]
        c_part = dcs_code[3:6]
        s_part = dcs_code[6:9]
        body = {k: v for k, v in {
            "originapplication": "RProPrismWeb",
            "active": 1,
            "sbssid": sbssid_val,
            "regional": False,
            "d": d_part,
            "c": c_part,
            "s": s_part,
            "dcscode": dcs_code,
            "dname": row.get("D_NAME") or "",
            "cname": row.get("C_NAME") or "",
            "sname": row.get("S_NAME") or "",
        }.items() if v is not None}
        resp = await http.post(
            f"{base_url}/api/backoffice/dcs",
            json={"data": [body]},
            headers={"Auth-Session": auth_session, "accept": "application/json, version=2",
                     "Content-Type": "application/json"},
        )
        try:
            resp_json = resp.json()
        except Exception:
            resp_json = {"raw": resp.text}
        if debug is not None:
            debug.update({
                "api_payload": {"data": [body]},
                "api_status": resp.status_code,
                "api_response": resp_json,
            })
        api_data = (resp_json.get("data") or []) if isinstance(resp_json, dict) else []
        sid = api_data[0].get("sid") if api_data else None
        if not sid:
            # Retry Oracle after create
            sid = await _oracle_scalar(
                f"SELECT sid FROM rps.dcs WHERE dcs_code = '{dcs_code}'", oc
            )
            if debug is not None:
                debug["oracle_retry_sid"] = sid

    if debug is not None:
        debug["final_sid"] = sid
    dcs_cache[dcs_code] = sid
    return sid


async def _vendsid(
    vend_code: str, row: dict,
    vend_cache: dict, sbs_cache: dict,
    oc: dict, http: httpx.AsyncClient,
    base_url: str, auth_session: str,
    debug: Optional[dict] = None,
) -> Optional[str]:
    if not vend_code:
        if debug is not None:
            debug.update({"vend_code": "(empty)", "skipped": "VEND_CODE column missing or empty"})
        return None
    if vend_code in vend_cache:
        if debug is not None:
            debug.update({"vend_code": vend_code, "source": "cache", "final_sid": vend_cache[vend_code]})
        return vend_cache[vend_code]

    sid = await _oracle_scalar(
        f"SELECT sid FROM rps.vendor WHERE vend_code = '{vend_code}'", oc
    )
    if debug is not None:
        debug.update({"vend_code": vend_code, "oracle_found": bool(sid), "oracle_sid": sid})

    if not sid:
        sbssid_val = await _sbssid(row.get("SBS_NO", "1"), sbs_cache, oc)
        body = {k: v for k, v in {
            "originapplication": "RProPrismWeb",
            "sbssid": sbssid_val,
            "regional": False,
            "vendcode": vend_code,
            "vendname": row.get("VEND_NAME") or vend_code,
            "active": True,
        }.items() if v is not None}
        resp = await http.post(
            f"{base_url}/api/backoffice/vendor",
            json={"data": [body]},
            headers={"Auth-Session": auth_session, "accept": "application/json, version=2",
                     "Content-Type": "application/json"},
        )
        try:
            resp_json = resp.json()
        except Exception:
            resp_json = {"raw": resp.text}
        if debug is not None:
            debug.update({
                "api_payload": {"data": [body]},
                "api_status": resp.status_code,
                "api_response": resp_json,
            })
        api_data = (resp_json.get("data") or []) if isinstance(resp_json, dict) else []
        sid = api_data[0].get("sid") if api_data else None
        if not sid:
            sid = await _oracle_scalar(
                f"SELECT sid FROM rps.vendor WHERE vend_code = '{vend_code}'", oc
            )
            if debug is not None:
                debug["oracle_retry_sid"] = sid

    if debug is not None:
        debug["final_sid"] = sid
    vend_cache[vend_code] = sid
    return sid


async def _taxcodesid(tax_code: str, tax_cache: dict, oc: dict) -> Optional[str]:
    if not tax_code:
        return None
    if tax_code not in tax_cache:
        tax_cache[tax_code] = await _oracle_scalar(
            f"SELECT sid FROM rps.tax_code WHERE tax_code = '{tax_code}'", oc
        )
    return tax_cache[tax_code]


async def _pref_reason_sid(cache: dict, oc: dict) -> Optional[str]:
    """Return the SID for the 'MANUALLY' preferred reason — cached for the whole batch."""
    key = "__MANUALLY__"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            "SELECT sid FROM RPS.PREF_REASON WHERE name = 'MANUALLY'", oc
        )
    return cache[key]


async def _store_sid(store_no: Any, sbssid: Optional[str], cache: dict, oc: dict) -> Optional[str]:
    """Return the store SID for the given store_no + subsidiary SID pair."""
    sn = str(store_no).strip() if store_no else "1"
    if not sbssid:
        return None
    key = f"{sn}::{sbssid}"
    if key not in cache:
        cache[key] = await _oracle_scalar(
            f"SELECT sid FROM rps.store WHERE store_no = {sn} AND sbs_sid = '{sbssid}'", oc
        )
    return cache[key]


# ── Payload builder ─────────────────────────────────────────────────────────────

def _to_json(v: Any) -> Any:
    import re as _re
    from datetime import datetime as _dt
    if v is None:
        return None
    if isinstance(v, _dt):
        return v.isoformat()
    # Float whole-numbers (e.g. 1e12 read from Excel) → "1000000000000"
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return v
    # String scientific-notation (e.g. "1E+12") → "1000000000000"
    if isinstance(v, str):
        stripped = v.strip()
        if _re.match(r'^-?\d+\.?\d*[eE][+\-]?\d+$', stripped):
            try:
                num = float(stripped)
                if num == int(num):
                    return str(int(num))
                return str(num)
            except (ValueError, OverflowError):
                pass
    return v


def build_payload(
    row: dict,
    sid_overrides: dict,
    existing_item: Optional[dict],
    pref_reason_sid: Optional[str] = None,
    store_sid: Optional[str] = None,
) -> dict:
    """
    Build the RetailPro InventorySaveItems payload for one row.

    The outer shape expected by the API is:
        {"data": [<this return value>]}

    where this return value is:
        {
            "OriginApplication": "RProPrismWeb",
            "PrimaryItemDefinition": { dcssid, vendsid, description1, … , sid },
            "InventoryItems": [ { …all fields…, "invnextend": […] } ],
            "SavingStyle": false,
            "localdate": "YYYY-MM-DDTHH:MM:SS",
            "DefaultReasonSidForQtyMemo":   <pref_reason_sid>,
            "DefaultReasonSidForCostMemo":  <pref_reason_sid>,
            "DefaultReasonSidForPriceMemo": <pref_reason_sid>,
            …
        }

    sid_overrides:    {json_field: value} – computed SIDs (dcssid, vendsid, etc.)
    existing_item:    API response item dict for update; None for create.
    pref_reason_sid:  SID from RPS.PREF_REASON WHERE name = 'MANUALLY'
    store_sid:        SID from rps.store for the row's store_no + subsidiary
    """
    # ── InventoryItems entry ──────────────────────────────────────────────────
    inv_item: dict = {}

    if existing_item:
        inv_item["sid"] = existing_item.get("sid")
    else:
        inv_item["sid"] = None

    # Map Excel columns → JSON fields (skip None and empty strings)
    for col, json_key in MAIN_FIELD_MAP.items():
        if col in row and row[col] is not None and str(row[col]).strip() != "" and json_key not in inv_item:
            inv_item[json_key] = _to_json(row[col])

    # Inject computed SIDs (overwrite whatever came from Excel)
    inv_item.update(sid_overrides)

    # These fields must always be doubles (float), not strings
    for _float_key in ("cost", "maxdiscperc1", "maxdiscperc2"):
        if _float_key in inv_item and inv_item[_float_key] is not None:
            try:
                inv_item[_float_key] = float(inv_item[_float_key])
            except (ValueError, TypeError):
                pass

    # These fields must always be integers, not strings
    for _int_key in ("useqtydecimals", "qtypercase"):
        if _int_key in inv_item and inv_item[_int_key] is not None:
            try:
                inv_item[_int_key] = int(float(inv_item[_int_key]))
            except (ValueError, TypeError):
                pass

    # regional and active: 0 → false, 1 → true
    for _bool_key in ("regional", "active"):
        if _bool_key in inv_item and inv_item[_bool_key] is not None:
            try:
                inv_item[_bool_key] = bool(int(float(inv_item[_bool_key])))
            except (ValueError, TypeError):
                pass

    # ── invnextend sub-object ─────────────────────────────────────────────────
    extend: dict = {}
    if existing_item:
        ex_ext = (existing_item.get("invnextend") or [{}])[0]
        if ex_ext.get("sid") is not None:
            extend["sid"] = ex_ext["sid"]
        if ex_ext.get("rowversion") is not None:
            extend["rowversion"] = ex_ext["rowversion"]
        if existing_item.get("sid") is not None:
            extend["invnsbsitemsid"] = existing_item["sid"]

    for col, json_key in INVNEXTEND_MAP.items():
        if col in row and row[col] is not None and str(row[col]).strip() != "":
            extend[json_key] = _to_json(row[col])

    inv_item["invnextend"] = [extend]

    # activestoresid — resolved from Oracle (store_no + subsidiary)
    if store_sid:
        inv_item["activestoresid"] = store_sid

    # ── PrimaryItemDefinition ─────────────────────────────────────────────────
    # Key order: sid → dcssid → vendsid → description1 → description2 → attribute → itemsize
    # dcssid and vendsid are always included (can be null, same as sid).
    primary_def: dict = {
        "sid":    existing_item.get("stylesid") if existing_item else None,
        "dcssid": sid_overrides.get("dcssid"),
        "vendsid": sid_overrides.get("vendsid"),
    }
    for k, v in {
        "description1": inv_item.get("description1"),
        "description2": inv_item.get("description2"),
        "attribute":    inv_item.get("attribute"),
        "itemsize":     inv_item.get("itemsize"),
    }.items():
        if v is not None and str(v).strip() != "":
            primary_def[k] = v

    # ── Outer wrapper (OriginApplication + PrimaryItemDefinition always first) ─
    outer: dict = {
        "OriginApplication":     "RProPrismWeb",
        "PrimaryItemDefinition": primary_def,
        "InventoryItems":        [inv_item],
        "UpdateStyleDefinition": False,
        "UpdateStyleCost":       False,
        "UpdateStylePrice":      False,
    }
    if pref_reason_sid:
        outer["DefaultReasonSidForQtyMemo"]   = pref_reason_sid
        outer["DefaultReasonSidForCostMemo"]  = pref_reason_sid
        outer["DefaultReasonSidForPriceMemo"] = pref_reason_sid
    return outer


# ── Per-row processor ───────────────────────────────────────────────────────────

async def process_row(
    row: dict,
    auth_session: str,
    oc: dict,
    base_url: str,
    http: httpx.AsyncClient,
    dcs_cache: dict,
    vend_cache: dict,
    tax_cache: dict,
    sbs_cache: dict,
    pref_reason_cache: dict,
    store_cache: dict,
) -> dict:
    upc = row.get("UPC", "")
    result: dict = {"upc": upc, "action": None, "sid": None, "ok": False, "error": None,
                    "description": row.get("DESCRIPTION1") or row.get("DESCRIPTION") or ""}
    dcs_debug: dict = {}
    vend_debug: dict = {}
    try:
        dcs_code  = str(row.get("DCS_CODE")  or "").strip()
        vend_code = str(row.get("VEND_CODE") or "").strip()
        tax_code  = str(row.get("TAX_CODE")  or "").strip()
        sbs_no    = str(row.get("SBS_NO")    or "1").strip()

        rp_headers = {
            "Auth-Session": auth_session,
            "accept": "application/json, version=2",
            "Content-Type": "application/json",
        }

        # ── 1. DCS ───────────────────────────────────────────────────────────
        dcssid_val = await _dcssid(
            dcs_code, row, dcs_cache, sbs_cache, oc, http, base_url, auth_session,
            debug=dcs_debug,
        )

        # ── 2. Vendor ────────────────────────────────────────────────────────
        vendsid_val = await _vendsid(
            vend_code, row, vend_cache, sbs_cache, oc, http, base_url, auth_session,
            debug=vend_debug,
        )

        # ── 3. Tax + subsidiary ──────────────────────────────────────────────
        taxcodesid_val = await _taxcodesid(tax_code, tax_cache, oc)
        sbssid_val     = await _sbssid(sbs_no, sbs_cache, oc)

        # ── 4. Preferred-reason SID (DefaultReasonSid* fields) ───────────────
        pref_reason_sid_val = await _pref_reason_sid(pref_reason_cache, oc)

        # ── 5. Store SID (activestoresid) ────────────────────────────────────
        store_no_val    = str(row.get("STORE_NO") or "1").strip()
        store_sid_val   = await _store_sid(store_no_val, sbssid_val, store_cache, oc)

        # ── 6. Check item by UPC ─────────────────────────────────────────────
        check_resp = await http.get(
            f"{base_url}/api/backoffice/inventory",
            params={"filter": f'(upc,eq,"{upc}")', "cols": "*,invnextend.*"},
            headers=rp_headers,
        )
        check_data = check_resp.json().get("data") or []

        # dcssid and vendsid are always included (null is valid and expected in both
        # PrimaryItemDefinition and InventoryItems).  taxcodesid / sbssid are only
        # included when resolved because sbssid has a per-update fallback below.
        sid_overrides: dict = {
            "dcssid":  dcssid_val,   # always present, may be None
            "vendsid": vendsid_val,  # always present, may be None
        }
        if taxcodesid_val is not None:
            sid_overrides["taxcodesid"] = taxcodesid_val
        if sbssid_val is not None:
            sid_overrides["sbssid"] = sbssid_val

        if len(check_data) > 0:
            existing = check_data[0]
            # For update, preserve existing sbssid if we couldn't resolve it
            if "sbssid" not in sid_overrides:
                sid_overrides["sbssid"] = existing.get("sbssid")
            payload = build_payload(
                row, sid_overrides, existing_item=existing,
                pref_reason_sid=pref_reason_sid_val,
                store_sid=store_sid_val,
            )
            action = "updated"
        else:
            payload = build_payload(
                row, sid_overrides, existing_item=None,
                pref_reason_sid=pref_reason_sid_val,
                store_sid=store_sid_val,
            )
            action = "created"

        # ── 5. Upsert ────────────────────────────────────────────────────────
        save_resp = await http.post(
            f"{base_url}/api/backoffice/inventory",
            params={"action": "InventorySaveItems"},
            json={"data": [payload]},
            headers=rp_headers,
        )
        save_json = save_resp.json()
        saved_data = save_json.get("data") or []
        # Response mirrors the new structure: data[0].InventoryItems[0].sid
        sid = None
        if saved_data:
            inv_items = saved_data[0].get("InventoryItems") or []
            if inv_items:
                sid = inv_items[0].get("sid")
            if not sid:
                sid = saved_data[0].get("sid")

        if save_resp.status_code in (200, 201) and sid:
            result.update({
                "action": action, "sid": sid, "ok": True,
                "api_response": save_json,
                "_dcs_debug": dcs_debug,
                "_vend_debug": vend_debug,
            })
        else:
            result.update({
                "action": action, "ok": False,
                "api_response": save_resp.text,
                "error": save_resp.text,
                "_payload_sent": {"data": [payload]},   # full request body for UI display
                "_dcs_debug": dcs_debug,
                "_vend_debug": vend_debug,
            })

    except Exception as exc:
        result["error"] = str(exc)
        result["_dcs_debug"] = dcs_debug
        result["_vend_debug"] = vend_debug

    return result


# ── DB persistence ──────────────────────────────────────────────────────────────

async def _persist_result(row: dict, result: dict, source_file: str) -> None:
    """Save a processed row result to the documents table."""
    import uuid as _uuid
    from app.db.postgres import get_session
    from app.models.document import Document

    ok = result.get("ok", False)
    error_msg = result.get("error")

    # Store the serialisable row dict (convert datetime → str)
    safe_row = {k: (_to_json(v)) for k, v in row.items()}

    import json as _json

    # On failure, embed the sent payload so the UI can show it alongside the error.
    # Stored as a JSON string (not a nested object) to preserve key insertion order
    # when retrieved from PostgreSQL JSONB.
    if not ok and result.get("_payload_sent"):
        safe_row["_payload_sent"] = _json.dumps(result["_payload_sent"], indent=2, ensure_ascii=False)

    # Always store DCS / vendor debug so the UI can show what happened to each SID.
    if result.get("_dcs_debug"):
        safe_row["_dcs_debug"] = _json.dumps(result["_dcs_debug"], indent=2, ensure_ascii=False)
    if result.get("_vend_debug"):
        safe_row["_vend_debug"] = _json.dumps(result["_vend_debug"], indent=2, ensure_ascii=False)

    async with get_session() as session:
        async with session.begin():
            doc = Document(
                id=_uuid.uuid4(),
                document_type="item_master",
                original_data=safe_row,
                retailprosid=result.get("sid") if ok else None,
                posted=ok,
                has_error=not ok,
                error_message=error_msg,          # full API response on failure
                source_file=source_file,
                posted_at=now_pkt() if ok else None,
            )
            session.add(doc)


# ── CSV parsing ─────────────────────────────────────────────────────────────────

def parse_csv_item_master(file_bytes: bytes) -> list[dict]:
    """
    Parse CSV bytes into item master row dicts — same format as parse_excel().
    Column headers are normalised to uppercase + stripped.
    Rows without a UPC value are skipped.
    Handles UTF-8 with or without BOM, and falls back to latin-1.
    """
    import csv as _csv
    import io as _io

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode CSV file (tried utf-8, latin-1).")

    reader = _csv.DictReader(_io.StringIO(text))
    rows_out: list[dict] = []
    for row in reader:
        normalised = {k.strip().upper(): v for k, v in row.items() if k}
        upc = str(normalised.get("UPC", "")).strip()
        if not upc:
            continue
        normalised["UPC"] = upc
        rows_out.append(normalised)
    return rows_out


# ── Shared batch runner ──────────────────────────────────────────────────────────

async def _run_rows_batch(rows: list[dict], source_file: str) -> dict:
    """
    Core pipeline shared by Excel and CSV paths:
      auth (once) → per-row processing with shared caches → persist each row.
    """
    from app.services.retailpro_auth import get_auth_session, sit_session, stand_session
    from app.db.settings_store import get_setting

    if not rows:
        return {"ok": False, "error": "No data rows found (all rows missing UPC).", "results": []}

    oc = {
        "host":         (await get_setting("oracle_host"))         or "",
        "port":         int((await get_setting("oracle_port"))     or "1521"),
        "service_name": (await get_setting("oracle_service_name")) or "",
        "username":     (await get_setting("oracle_username"))     or "",
        "password":     (await get_setting("oracle_password"))     or "",
    }

    base_url     = ((await get_setting("retailpro_base_url")) or "").rstrip("/")
    auth_session = await get_auth_session()

    # Activate the session immediately after obtaining it
    await sit_session(base_url, auth_session)

    dcs_cache:          dict = {}
    vend_cache:         dict = {}
    tax_cache:          dict = {}
    sbs_cache:          dict = {}
    pref_reason_cache:  dict = {}
    store_cache:        dict = {}

    results: list[dict] = []
    try:
        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as http:
            for row in rows:
                row_result = await process_row(
                    row=row,
                    auth_session=auth_session,
                    oc=oc,
                    base_url=base_url,
                    http=http,
                    dcs_cache=dcs_cache,
                    vend_cache=vend_cache,
                    tax_cache=tax_cache,
                    sbs_cache=sbs_cache,
                    pref_reason_cache=pref_reason_cache,
                    store_cache=store_cache,
                )
                results.append(row_result)
                try:
                    await _persist_result(row, row_result, source_file)
                except Exception as db_exc:
                    logger.warning("DB persist failed for UPC %s: %s", row_result["upc"], db_exc)
                logger.info(
                    "Item %s: %s %s",
                    row_result["upc"],
                    row_result["action"],
                    "✓" if row_result["ok"] else f"✗ {row_result['error']}",
                )
    finally:
        # Always destroy the session after all rows are processed (or on error)
        await stand_session(base_url, auth_session)

    created = sum(1 for r in results if r["ok"] and r["action"] == "created")
    updated = sum(1 for r in results if r["ok"] and r["action"] == "updated")
    errors  = sum(1 for r in results if not r["ok"])

    return {
        "ok":      True,
        "total":   len(results),
        "created": created,
        "updated": updated,
        "errors":  errors,
        "results": results,
    }


# ── Batch entry points ───────────────────────────────────────────────────────────

async def process_excel_batch(file_bytes: bytes, source_file: str = "item_master_import.xlsx") -> dict:
    """Parse an Excel (.xlsx) file and run the full item master pipeline."""
    rows = parse_excel(file_bytes)
    return await _run_rows_batch(rows, source_file)


async def process_csv_batch(file_bytes: bytes, source_file: str = "item_master_import.csv") -> dict:
    """Parse a CSV file and run the full item master pipeline."""
    rows = parse_csv_item_master(file_bytes)
    return await _run_rows_batch(rows, source_file)
